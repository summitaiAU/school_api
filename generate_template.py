import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def create_sample_excel_template(filename="school_data_template.xlsx"):
    """
    Create a sample Excel template for school data import
    """
    # Create a Pandas Excel writer using openpyxl as the engine
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    
    # Create sample data for Schools sheet
    schools_data = {
        'name': ['East High School', ''],
        'address': ['123 Main St, Cityville', ''],
        'phone': ['555-1234', ''],
        'email': ['info@easthigh.edu', '']
    }
    schools_df = pd.DataFrame(schools_data)
    
    # Create sample data for Classes sheet
    classes_data = {
        'school_name': ['East High School', ''],
        'name': ['Class 9A', ''],
        'grade_level': ['9', ''],
        'academic_year': ['2024-2025', '']
    }
    classes_df = pd.DataFrame(classes_data)
    
    # Create sample data for Students sheet
    students_data = {
        'first_name': ['John', ''],
        'last_name': ['Doe', ''],
        'student_id': ['S12345', ''],
        'date_of_birth': ['2008-05-12', ''],
        'school_name': ['East High School', ''],
        'class_name': ['Class 9A', ''],
        'email': ['john.doe@example.com', ''],
        'address': ['456 Oak Rd, Townsville', ''],
        'parent_name': ['Jane Doe', ''],
        'parent_contact': ['555-5678', '']
    }
    students_df = pd.DataFrame(students_data)
    
    # Write each DataFrame to a different worksheet
    schools_df.to_excel(writer, sheet_name='Schools', index=False)
    classes_df.to_excel(writer, sheet_name='Classes', index=False)
    students_df.to_excel(writer, sheet_name='Students', index=False)
    
    # Get the openpyxl workbook and sheets
    workbook = writer.book
    
    # Format the sheets
    for sheet_name in ['Schools', 'Classes', 'Students']:
        worksheet = workbook[sheet_name]
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        # Format headers
        for col_num, column_title in enumerate(worksheet[1], 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto adjust column width based on content
            column_letter = openpyxl.utils.get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].width = max(15, len(str(column_title.value)) + 2)
        
        # Format data cells
        for row in range(2, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
        
        # Add instructions as a comment to the first cell
        if sheet_name == 'Schools':
            instructions = "Enter school information. Each school must have a unique name."
        elif sheet_name == 'Classes':
            instructions = "Enter class information. The school_name must match exactly with a name in the Schools sheet."
        else:  # Students
            instructions = "Enter student information. The school_name and class_name must match exactly with existing records."
        
        first_cell = worksheet['A1']
        from openpyxl.comments import Comment
        comment = Comment(instructions, "Template Generator")
        first_cell.comment = comment
    
    # Add an Instructions sheet
    instructions_sheet = workbook.create_sheet(title="Instructions", index=0)
    
    instructions = [
        ["School Data Import Template Instructions"],
        [""],
        ["This template contains three sheets that must be filled in the following order:"],
        [""],
        ["1. Schools: Enter information about each school"],
        ["   - Each school must have a unique name"],
        [""],
        ["2. Classes: Enter information about classes"],
        ["   - The school_name must match exactly with a name in the Schools sheet"],
        [""],
        ["3. Students: Enter information about students"],
        ["   - The school_name must match a name in the Schools sheet"],
        ["   - The class_name must match a name in the Classes sheet for the given school"],
        ["   - The date_of_birth should be in YYYY-MM-DD format (e.g., 2008-05-12)"],
        [""],
        ["Important Notes:"],
        ["- Do not modify the column headers"],
        ["- Do not add additional sheets"],
        ["- Make sure to fill in all required fields (name, school_name, class_name, first_name, last_name)"]
    ]
    
    for row_num, row_data in enumerate(instructions, 1):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = instructions_sheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            if row_num == 1:
                cell.font = Font(bold=True, size=14)
            elif row_data and row_data[0].startswith("Important Notes:"):
                cell.font = Font(bold=True)
    
    # Adjust column width for Instructions sheet
    instructions_sheet.column_dimensions['A'].width = 60
    
    # Save the workbook
    writer.close()
    
    print(f"Template created successfully: {filename}")
    return filename

if __name__ == "__main__":
    create_sample_excel_template()